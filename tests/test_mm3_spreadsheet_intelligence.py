"""
CLAUDE-MM3 (Spreadsheet and Structured-Data Intelligence) tests:
CaseWorkspaceStore.register_spreadsheet_structure (already-extracted sheet
data -> governed MM1 evidence) and services/spreadsheet_intelligence's real
openpyxl/csv read, classification, and bounded-edit orchestration.

Unlike tests/test_mm2_pdf_document_intelligence.py (which mocks the pypdf
call site, matching this repo's own established PDF-testing convention),
this file builds REAL small .xlsx workbooks in-memory via openpyxl's own
writer for most tests - openpyxl is now a first-class dependency (see
requirements.txt), so there is no reason to fake its own output; a real
build-then-read round trip is strictly better evidence than a mock would
be, and openpyxl's writer is fast/deterministic enough that this adds no
real cost.

Run via:

    python -m unittest tests.test_mm3_spreadsheet_intelligence -v
"""
from __future__ import annotations

import io
import shutil
import tempfile
import unittest
from pathlib import Path

import openpyxl

from services.case_workspace import (
    EVIDENCE_CLASS_DIRECT_SOURCE,
    KNOWN_SPREADSHEET_CLASSIFICATIONS,
    OBSERVATION_AUTHOR_HUMAN,
    SPREADSHEET_CLASSIFICATION_ENCRYPTED_OR_UNSUPPORTED,
    SPREADSHEET_CLASSIFICATION_EXCESSIVE_SIZE,
    SPREADSHEET_CLASSIFICATION_MALFORMED,
    SPREADSHEET_CLASSIFICATION_SUPPORTED,
    CaseWorkspaceError,
    CaseWorkspaceStore,
    ConcurrentModificationError,
)
from services.governance import GovernanceLog
from services.spreadsheet_intelligence import (
    MAX_ROWS_PER_SHEET,
    MAX_UNCOMPRESSED_BYTES,
    SpreadsheetIntelligenceError,
    apply_bounded_cell_edit,
    inspect_workbook,
    register_spreadsheet_evidence_for_source,
    safe_csv_cell,
)


def _build_risk_register_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk Log"
    ws.append(["Risk ID", "Description", "Probability", "Impact", "Score"])
    ws.append(["R-001", "Foundation delay risk", 0.3, 5, "=C2*D2"])
    ws.append(["R-002", "Weather impact", 0.5, 3, "=C3*D3"])
    hidden = wb.create_sheet("Internal Notes")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "not for client eyes"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_workbook_with_table() -> bytes:
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Team"
    ws.append(["ID", "Name", "Role"])
    ws.append(["001", "Jane Doe", "PM"])
    ws.append(["002", "John Smith", "Engineer"])
    table = Table(displayName="TeamTable", ref="A1:C3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1")
    ws.add_table(table)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_workbook_with_external_link() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["a"])
    ws["B1"] = "=[1]Sheet1!A1"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class SpreadsheetStructureStoreTests(unittest.TestCase):
    """CaseWorkspaceStore.register_spreadsheet_structure - already-
    extracted sheet dicts in, governed StructuralUnit/AddressableRegion/
    EvidenceItem out. No openpyxl/csv involved at this layer."""

    def setUp(self):
        self.tmp_dir = Path(tempfile.mkdtemp(prefix="beehive_test_mm3_store_"))
        self.store = CaseWorkspaceStore(self.tmp_dir)
        self.gov = GovernanceLog(self.tmp_dir)
        self.project_id = "test-project-mm3"
        self.workspace = self.store.get_or_create(self.project_id)
        self.source = self.store.add_source(
            self.workspace, name="risk_register.xlsx", file_path="/tmp/risk_register.xlsx", kind="project_document",
        )
        self.other_workspace = self.store.get_or_create("test-project-mm3-other")
        self.other_source = self.store.add_source(
            self.other_workspace, name="other.xlsx", file_path="/tmp/other.xlsx", kind="project_document",
        )

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def _sample_sheets(self):
        return [
            {
                "name": "Risk Log", "index": 0, "visible": True, "row_count": 2, "column_count": 2,
                "truncated": False,
                "rows": [
                    {"row_index": 1, "cells": {
                        "A": {"value": "Risk ID", "formula": None, "cached_value": "Risk ID", "data_type": "s"},
                        "B": {"value": "Description", "formula": None, "cached_value": "Description", "data_type": "s"},
                    }},
                    {"row_index": 2, "cells": {
                        "A": {"value": "R-001", "formula": None, "cached_value": "R-001", "data_type": "s"},
                        "B": {"value": None, "formula": "=A2&\"!\"", "cached_value": "R-001!", "data_type": "f"},
                    }},
                ],
            },
            {
                "name": "Notes", "index": 1, "visible": False, "row_count": 0, "column_count": 0,
                "truncated": False, "rows": [],
            },
        ]

    def test_worksheet_structural_units_and_sheet_order(self):
        result = self.store.register_spreadsheet_structure(self.workspace, self.source["id"], self._sample_sheets())
        self.assertEqual(len(result["structural_unit_ids"]), 2)
        units = self.store.structural_units_for_source(self.workspace, self.source["id"])
        ordered = sorted(units, key=lambda u: u["order_index"])
        self.assertEqual([u["label"] for u in ordered], ["Risk Log", "Notes"])
        self.assertEqual(ordered[0]["unit_type"], "worksheet")

    def test_hidden_sheet_still_gets_a_real_structural_unit(self):
        result = self.store.register_spreadsheet_structure(self.workspace, self.source["id"], self._sample_sheets())
        units = self.store.structural_units_for_source(self.workspace, self.source["id"])
        notes_unit = next(u for u in units if u["label"] == "Notes")
        self.assertFalse(notes_unit["modality_metadata"]["visible"])
        self.assertIn(notes_unit["id"], result["structural_unit_ids"])

    def test_cell_and_row_identity(self):
        result = self.store.register_spreadsheet_structure(self.workspace, self.source["id"], self._sample_sheets())
        self.assertEqual(len(result["addressable_region_ids"]), 2)
        regions = [self.store.get_addressable_region(self.workspace, rid) for rid in result["addressable_region_ids"]]
        self.assertEqual({r["address"]["row_index"] for r in regions}, {1, 2})
        for r in regions:
            self.assertEqual(r["region_type"], "row")
            self.assertEqual(r["address"]["sheet_name"], "Risk Log")

    def test_value_type_and_formula_distinction_preserved_in_region(self):
        result = self.store.register_spreadsheet_structure(self.workspace, self.source["id"], self._sample_sheets())
        regions = [self.store.get_addressable_region(self.workspace, rid) for rid in result["addressable_region_ids"]]
        row2 = next(r for r in regions if r["address"]["row_index"] == 2)
        cell_b = row2["address"]["cells"]["B"]
        self.assertEqual(cell_b["data_type"], "f")
        self.assertEqual(cell_b["formula"], '=A2&"!"')
        self.assertEqual(cell_b["cached_value"], "R-001!")
        self.assertIsNone(cell_b["value"])

    def test_single_cell_region(self):
        unit = self.store.create_structural_unit(self.workspace, self.source["id"], "worksheet", 0, label="Risk Log")
        region = self.store.create_addressable_cell_region(
            self.workspace, unit["id"], sheet_name="Risk Log", cell_ref="B2", value="Foundation delay risk",
        )
        self.assertEqual(region["region_type"], "cell")
        self.assertEqual(region["address"]["cell_ref"], "B2")

    def test_citation_rendering(self):
        result = self.store.register_spreadsheet_structure(self.workspace, self.source["id"], self._sample_sheets())
        region_id = result["addressable_region_ids"][0]
        citation = self.store.resolve_region_citation(self.workspace, region_id)
        self.assertEqual(citation["status"], "resolved")
        self.assertIn("risk_register.xlsx", citation["label"])
        self.assertIn("Risk Log", citation["label"])
        self.assertIn("row 1", citation["label"])

    def test_project_isolation_and_cross_project_denial(self):
        self.store.register_spreadsheet_structure(self.workspace, self.source["id"], self._sample_sheets())
        self.store.register_spreadsheet_structure(self.other_workspace, self.other_source["id"], self._sample_sheets())
        self.assertEqual(len(self.workspace.structural_units), 2)
        self.assertEqual(len(self.other_workspace.structural_units), 2)
        with self.assertRaises(CaseWorkspaceError):
            self.store.register_spreadsheet_structure(self.workspace, self.other_source["id"], self._sample_sheets())

    def test_falsification_stale_broken_anchor(self):
        result = self.store.register_spreadsheet_structure(self.workspace, self.source["id"], self._sample_sheets())
        region_id = result["addressable_region_ids"][0]
        self.assertEqual(self.store.resolve_region_citation(self.workspace, region_id)["status"], "resolved")
        live_source = self.store._find(self.workspace.sources, self.source["id"])
        live_source["removed_at"] = "2026-08-06T00:00:00+00:00"
        self.store.save(self.workspace)
        self.assertEqual(self.store.resolve_region_citation(self.workspace, region_id)["status"], "unavailable")
        self.assertEqual(self.store.resolve_region_citation(self.workspace, "does-not-exist")["status"], "unavailable")

    def test_persistence_round_trip(self):
        result = self.store.register_spreadsheet_structure(self.workspace, self.source["id"], self._sample_sheets())
        reloaded = self.store.get(self.project_id)
        self.assertEqual(len(reloaded.structural_units), 2)
        self.assertEqual(reloaded.structural_units[0]["id"], result["structural_unit_ids"][0])

    def test_backward_compatibility_legacy_workspace(self):
        import json
        path = self.store._path_for(self.project_id)
        data = json.loads(path.read_text(encoding="utf-8"))
        for key in ("structural_units", "addressable_regions", "evidence_items"):
            data.pop(key, None)
        path.write_text(json.dumps(data), encoding="utf-8")
        reloaded = self.store.get(self.project_id)
        result = self.store.register_spreadsheet_structure(reloaded, self.source["id"], self._sample_sheets())
        self.assertTrue(result["structural_unit_ids"])

    def test_concurrent_modification_is_detected(self):
        copy_a = self.store.get(self.project_id)
        copy_b = self.store.get(self.project_id)
        self.store.register_spreadsheet_structure(copy_a, self.source["id"], self._sample_sheets())
        with self.assertRaises(ConcurrentModificationError):
            self.store.register_spreadsheet_structure(copy_b, self.source["id"], self._sample_sheets())

    def test_direct_evidence_vs_derived_observation(self):
        result = self.store.register_spreadsheet_structure(self.workspace, self.source["id"], self._sample_sheets())
        evidence_id = result["evidence_item_ids"][0]
        observation = self.store.record_derived_observation(
            self.workspace, statement="This register shows an unmitigated foundation risk.",
            author_type=OBSERVATION_AUTHOR_HUMAN, author="tester", method="review",
            supporting_evidence_ids=[evidence_id],
        )
        self.assertIn(evidence_id, observation["supporting_evidence_ids"])
        evidence = self.store.get_evidence_item(self.workspace, evidence_id)
        self.assertEqual(evidence["evidence_class"], EVIDENCE_CLASS_DIRECT_SOURCE)
        self.assertNotIn("author_type", evidence)


class SpreadsheetIntelligenceOrchestrationTests(unittest.TestCase):
    """Real openpyxl/csv reads - classification, security detections,
    bounded editing. No mocking - openpyxl is a first-class dependency."""

    def setUp(self):
        self.tmp_dir = Path(tempfile.mkdtemp(prefix="beehive_test_mm3_orch_"))
        self.store = CaseWorkspaceStore(self.tmp_dir)
        self.gov = GovernanceLog(self.tmp_dir)
        self.project_id = "test-project-mm3-orch"
        self.workspace = self.store.get_or_create(self.project_id)

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def _add_xlsx_source(self, workspace, filename, content_bytes):
        path = self.tmp_dir / filename
        path.write_bytes(content_bytes)
        return self.store.add_source(workspace, name=filename, file_path=str(path), kind="project_document")

    # -- classification -----------------------------------------------------

    def test_workbook_classification_supported(self):
        source = self._add_xlsx_source(self.workspace, "risk.xlsx", _build_risk_register_xlsx())
        result = register_spreadsheet_evidence_for_source(self.store, self.workspace, source["id"], actor="tester")
        self.assertEqual(result["classification"], SPREADSHEET_CLASSIFICATION_SUPPORTED)
        self.assertEqual(set(KNOWN_SPREADSHEET_CLASSIFICATIONS), {
            SPREADSHEET_CLASSIFICATION_SUPPORTED, SPREADSHEET_CLASSIFICATION_MALFORMED,
            SPREADSHEET_CLASSIFICATION_ENCRYPTED_OR_UNSUPPORTED, SPREADSHEET_CLASSIFICATION_EXCESSIVE_SIZE,
        })

    def test_malformed_workbook_classification(self):
        source = self._add_xlsx_source(self.workspace, "bad.xlsx", b"not a real xlsx file")
        result = register_spreadsheet_evidence_for_source(self.store, self.workspace, source["id"])
        self.assertEqual(result["classification"], SPREADSHEET_CLASSIFICATION_MALFORMED)
        reloaded = self.store.get(self.project_id)
        self.assertEqual(reloaded.structural_units, [])

    def test_encrypted_workbook_classification(self):
        source = self._add_xlsx_source(self.workspace, "enc.xlsx", b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64)
        result = register_spreadsheet_evidence_for_source(self.store, self.workspace, source["id"])
        self.assertEqual(result["classification"], SPREADSHEET_CLASSIFICATION_ENCRYPTED_OR_UNSUPPORTED)

    def test_excessive_size_classification(self):
        insp = inspect_workbook(b"x" * (MAX_UNCOMPRESSED_BYTES + 1), "huge.xlsx")
        self.assertEqual(insp["classification"], SPREADSHEET_CLASSIFICATION_EXCESSIVE_SIZE)

    def test_macro_enabled_content_is_detected_even_if_renamed_xlsx(self):
        """Defense in depth: a renamed .xlsm masquerading as .xlsx (real
        VBA project inside the zip) is still refused, never executed."""
        wb = openpyxl.Workbook()
        wb.active.append(["a"])
        buf = io.BytesIO()
        wb.save(buf)
        import zipfile
        src_zip = zipfile.ZipFile(io.BytesIO(buf.getvalue()))
        out_buf = io.BytesIO()
        with zipfile.ZipFile(out_buf, "w") as out_zip:
            for item in src_zip.infolist():
                out_zip.writestr(item, src_zip.read(item.filename))
            out_zip.writestr("xl/vbaProject.bin", b"fake macro bytes")
        insp = inspect_workbook(out_buf.getvalue(), "renamed.xlsx")
        self.assertEqual(insp["classification"], SPREADSHEET_CLASSIFICATION_ENCRYPTED_OR_UNSUPPORTED)
        self.assertTrue(any("macro" in w.lower() for w in insp["warnings"]))

    # -- worksheet structure / hidden sheets / tables / external links ----------

    def test_worksheet_names_order_and_hidden_detection(self):
        insp = inspect_workbook(_build_risk_register_xlsx(), "risk.xlsx")
        names = [s["name"] for s in insp["sheets"]]
        self.assertEqual(names, ["Risk Log", "Internal Notes"])
        hidden_sheet = insp["sheets"][1]
        self.assertFalse(hidden_sheet["visible"])
        self.assertTrue(any("hidden" in w for w in insp["warnings"]))

    def test_named_table_detection(self):
        insp = inspect_workbook(_build_workbook_with_table(), "team.xlsx")
        self.assertTrue(any("table" in w.lower() for w in insp["warnings"]))

    def test_external_link_detection(self):
        """
        A genuine external-link reference (`=[1]Sheet1!A1`, Excel's own
        syntax for "another workbook") is written into the formula, but
        openpyxl's writer only round-trips a real `_external_links` entry
        when the referenced workbook was itself opened alongside this one
        - not practical to construct through the public API alone. Proven
        instead the way this codebase already proves an adjacent case
        (drawing_intake.py's own DrawingIntakeResult tests mock at the
        library boundary, not by hand-building exotic real files): a real
        workbook is loaded for real, then `inspect_workbook`'s own
        detection line is exercised directly against a workbook object
        carrying a real-shaped `_external_links` list, proving the
        DETECTION logic itself (not openpyxl's writer) is what this test
        is actually about.
        """
        from unittest.mock import patch

        real_bytes = _build_workbook_with_external_link()
        original_load = openpyxl.load_workbook

        def _load_with_fake_external_link(*args, **kwargs):
            wb = original_load(*args, **kwargs)
            if not kwargs.get("data_only"):
                wb._external_links = [object()]
            return wb

        with patch("services.spreadsheet_intelligence.openpyxl.load_workbook", side_effect=_load_with_fake_external_link):
            insp = inspect_workbook(real_bytes, "linked.xlsx")
        self.assertTrue(any("external link" in w for w in insp["warnings"]))

    def test_formula_preservation_and_cached_result_distinction(self):
        insp = inspect_workbook(_build_risk_register_xlsx(), "risk.xlsx")
        risk_sheet = insp["sheets"][0]
        row2 = next(r for r in risk_sheet["rows"] if r["row_index"] == 2)
        formula_cell = row2["cells"]["E"]
        self.assertEqual(formula_cell["formula"], "=C2*D2")
        self.assertIsNone(formula_cell["value"])
        # Built programmatically, never opened by real Excel - honestly
        # None, never fabricated as a freshly-computed result.
        self.assertIsNone(formula_cell["cached_value"])
        self.assertTrue(any("not recalculated" in w for w in insp["warnings"]))

    # -- editing / provenance / concurrency --------------------------------------

    def test_edit_provenance_before_after_and_formula_preserved(self):
        source = self._add_xlsx_source(self.workspace, "risk.xlsx", _build_risk_register_xlsx())
        register_spreadsheet_evidence_for_source(self.store, self.workspace, source["id"])
        result = apply_bounded_cell_edit(
            self.store, self.workspace, source["id"], sheet_name="Risk Log", cell_ref="B2",
            new_value="Foundation delay risk (revised)", actor="tester", governance_log=self.gov,
        )
        self.assertEqual(result["before_value"], "Foundation delay risk")
        self.assertEqual(result["after_value"], "Foundation delay risk (revised)")
        self.assertTrue(Path(result["backup_path"]).exists())

        reopened = openpyxl.load_workbook(self.tmp_dir / "risk.xlsx", data_only=False)
        ws = reopened["Risk Log"]
        self.assertEqual(ws["B2"].value, "Foundation delay risk (revised)")
        self.assertEqual(ws["E2"].value, "=C2*D2")  # formula untouched
        self.assertEqual(ws["E2"].data_type, "f")

        events = self.gov.read(self.project_id)
        edit_events = [e for e in events if e.event_type == "spreadsheet_cell_edited"]
        self.assertEqual(len(edit_events), 1)
        self.assertEqual(edit_events[0].payload["before_value"], "Foundation delay risk")

    def test_refuses_editing_a_formula_cell(self):
        source = self._add_xlsx_source(self.workspace, "risk.xlsx", _build_risk_register_xlsx())
        with self.assertRaises(SpreadsheetIntelligenceError):
            apply_bounded_cell_edit(self.store, self.workspace, source["id"], "Risk Log", "E2", "999")

    def test_falsification_concurrency_guard_is_real(self):
        """Prove expected_file_hash is load-bearing: a stale hash is
        rejected; the correct current hash succeeds."""
        source = self._add_xlsx_source(self.workspace, "risk.xlsx", _build_risk_register_xlsx())
        with self.assertRaises(SpreadsheetIntelligenceError):
            apply_bounded_cell_edit(
                self.store, self.workspace, source["id"], "Risk Log", "B2", "x",
                expected_file_hash="0" * 64,
            )
        import hashlib
        real_hash = hashlib.sha256((self.tmp_dir / "risk.xlsx").read_bytes()).hexdigest()
        result = apply_bounded_cell_edit(
            self.store, self.workspace, source["id"], "Risk Log", "B2", "x", expected_file_hash=real_hash,
        )
        self.assertEqual(result["after_value"], "x")

    def test_leading_zero_identifier_preserved_on_edit(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Team"
        ws.append(["ID", "Name"])
        ws["A2"] = "001"
        buf = io.BytesIO()
        wb.save(buf)
        source = self._add_xlsx_source(self.workspace, "team.xlsx", buf.getvalue())
        apply_bounded_cell_edit(self.store, self.workspace, source["id"], "Team", "B2", "Jane Doe")
        reopened = openpyxl.load_workbook(self.tmp_dir / "team.xlsx")
        self.assertEqual(reopened["Team"]["A2"].value, "001")  # untouched cell, leading zero intact

    def test_backup_preserves_original_before_edit(self):
        source = self._add_xlsx_source(self.workspace, "risk.xlsx", _build_risk_register_xlsx())
        result = apply_bounded_cell_edit(self.store, self.workspace, source["id"], "Risk Log", "B2", "changed")
        backup_wb = openpyxl.load_workbook(result["backup_path"])
        self.assertEqual(backup_wb["Risk Log"]["B2"].value, "Foundation delay risk")

    def test_edit_rejects_non_xlsx_source(self):
        source = self._add_xlsx_source(self.workspace, "notes.txt", b"hello")
        with self.assertRaises(SpreadsheetIntelligenceError):
            apply_bounded_cell_edit(self.store, self.workspace, source["id"], "Sheet1", "A1", "x")

    def test_edit_rejects_cross_project_source(self):
        other_workspace = self.store.get_or_create("test-project-mm3-orch-other")
        source = self._add_xlsx_source(other_workspace, "risk.xlsx", _build_risk_register_xlsx())
        with self.assertRaises(SpreadsheetIntelligenceError):
            apply_bounded_cell_edit(self.store, self.workspace, source["id"], "Risk Log", "B2", "x")

    # -- CSV -----------------------------------------------------------------

    def test_csv_leading_zero_preserved(self):
        source = self._add_xlsx_source(
            self.workspace, "team.csv", b"Team ID,Name,Role\n001,Jane Doe,PM\n002,John Smith,Engineer\n",
        )
        result = register_spreadsheet_evidence_for_source(self.store, self.workspace, source["id"], actor="tester")
        self.assertEqual(result["classification"], SPREADSHEET_CLASSIFICATION_SUPPORTED)
        reloaded = self.store.get(self.project_id)
        unit = self.store.structural_units_for_source(reloaded, source["id"])[0]
        rows = self.store.regions_for_structural_unit(reloaded, unit["id"])
        data_row = next(r for r in rows if r["address"]["row_index"] == 2)
        self.assertEqual(data_row["address"]["cells"]["A"]["value"], "001")

    def test_csv_formula_injection_export_safety(self):
        self.assertEqual(safe_csv_cell("=SUM(A1:A2)"), "'=SUM(A1:A2)")
        self.assertEqual(safe_csv_cell("+1+1"), "'+1+1")
        self.assertEqual(safe_csv_cell("-1"), "'-1")
        self.assertEqual(safe_csv_cell("@cmd"), "'@cmd")
        self.assertEqual(safe_csv_cell("normal text"), "normal text")
        self.assertEqual(safe_csv_cell("001"), "001")
        self.assertEqual(safe_csv_cell(None), "")

    def test_backward_compatibility_existing_csv_ingestion_unaffected(self):
        """Section 20's own explicit checklist item: the pre-existing,
        unrelated CSV ingestion path (BHiveParser treating .csv as plain
        text for the requirement-extraction pipeline) must still work,
        completely untouched by this module."""
        from services.bhive_parser import BHiveParser

        parser = BHiveParser(anthropic_api_key=None)
        parsed = parser.parse(b"Section 1\nThe contractor shall provide labor.\n", "spec.csv")
        self.assertEqual(parsed.filename, "spec.csv")


class SpreadsheetApiRetrievalTests(unittest.TestCase):
    """Functional (not just auth) proof the new /api/v1 routes work with
    real registered data - mirrors MM1/MM2's own MM1ApiRetrievalTests/
    MM2 API coverage."""

    def setUp(self):
        import app as app_module
        from services.bhive_parser import ParsedDocument
        from services.requirements_registry import RequirementsRegistry

        self.tmp_dir = Path(tempfile.mkdtemp(prefix="beehive_test_mm3_api_"))
        self.flask_app = app_module.create_app("testing")
        self.flask_app.config["REGISTRY_STORE_PATH"] = str(self.tmp_dir)
        self.project_id = "mm3-api-project"

        document = ParsedDocument(
            project_id=self.project_id, filename="spec.txt", ingested_at="2026-01-01T00:00:00+00:00",
        )
        RequirementsRegistry(self.tmp_dir).save(document)

        self.store = CaseWorkspaceStore(self.tmp_dir)
        workspace = self.store.get_or_create(self.project_id)
        workspace.owner = "tester"
        self.store.save(workspace)
        xlsx_path = self.tmp_dir / "risk.xlsx"
        xlsx_path.write_bytes(_build_risk_register_xlsx())
        self.source = self.store.add_source(
            workspace, name="risk.xlsx", file_path=str(xlsx_path), kind="project_document",
        )

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def _client_as_admin(self):
        client = self.flask_app.test_client()
        with client.session_transaction() as sess:
            sess["user_id"] = 1
            sess["username"] = "tester"
            sess["role"] = "admin"
        return client

    def test_register_structure_then_edit_via_api(self):
        client = self._client_as_admin()
        response = client.post(f"/api/v1/documents/{self.project_id}/sources/{self.source['id']}/spreadsheet-structure")
        self.assertEqual(response.status_code, 201)
        body = response.get_json()
        self.assertEqual(body["classification"], SPREADSHEET_CLASSIFICATION_SUPPORTED)

        edit_response = client.post(
            f"/api/v1/documents/{self.project_id}/sources/{self.source['id']}/spreadsheet-cell",
            json={"sheet_name": "Risk Log", "cell_ref": "B2", "value": "revised via API"},
        )
        self.assertEqual(edit_response.status_code, 200)
        self.assertEqual(edit_response.get_json()["after_value"], "revised via API")

    def test_edit_missing_fields_returns_400(self):
        client = self._client_as_admin()
        response = client.post(
            f"/api/v1/documents/{self.project_id}/sources/{self.source['id']}/spreadsheet-cell", json={},
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.get_json()["error"], "invalid_edit")


if __name__ == "__main__":
    unittest.main()
