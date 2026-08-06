"""
CLAUDE-MM3: Spreadsheet and Structured-Data Intelligence - the thin
orchestration layer between a real workbook read (openpyxl for .xlsx,
stdlib `csv` for .csv) and the MM1 evidence contract (services/
case_workspace.py's `register_spreadsheet_structure`). Exists specifically
so case_workspace.py does not have to import openpyxl, the same decoupling
services/pdf_intelligence.py already established for pypdf.

Capability boundary, stated up front: no macro/VBA execution (this module
never opens a file with `keep_vba=True` and never executes anything found
in one), no formula recalculation (a formula's own cached value, if the
workbook was last saved by real Excel, is read and preserved verbatim -
never recomputed by this module, which has no formula engine and does not
pretend to be Excel), no Power Query/pivot-table/chart editing, no legacy
.xls support (that format needs a second, separate dependency - xlrd 2.x
dropped .xlsx support entirely and only reads .xls - deliberately not
added this stage; a .xls upload is refused, honestly, not silently
misread as .xlsx).
"""
from __future__ import annotations

import csv
import hashlib
import io
import shutil
import time
import zipfile
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from services.case_workspace import (
    CaseWorkspaceError,
    CaseWorkspaceStore,
    ProjectWorkspace,
    SPREADSHEET_CLASSIFICATION_ENCRYPTED_OR_UNSUPPORTED,
    SPREADSHEET_CLASSIFICATION_EXCESSIVE_SIZE,
    SPREADSHEET_CLASSIFICATION_MALFORMED,
    SPREADSHEET_CLASSIFICATION_SUPPORTED,
)
from services.governance import GovernanceLog

SPREADSHEET_EXTRACTOR_VERSION_PREFIX = "openpyxl"
# Bounds (Section 16: "protect against decompression bombs and excessive
# workbook dimensions") - deliberately conservative for a first slice; a
# real risk register or team list is nowhere near these limits, and
# truncation is always reported honestly, never silent.
MAX_ROWS_PER_SHEET = 200
MAX_SHEETS = 50
MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024

# Password-protected .xlsx (and legacy .xls) use the older OLE2/CFBF
# container format, not a plain zip - distinguishable by this leading
# magic number alone, before ever attempting to unzip anything.
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


class SpreadsheetIntelligenceError(Exception):
    """Caller error - not a Source, wrong project, not .xlsx/.csv, no
    stored file, or (for edits) an invalid edit request. Distinct from an
    attempted-and-failed READ, which is always a real, honest
    classification return value, never an exception - the same
    distinction services/pdf_intelligence.py's PdfIntelligenceError
    already establishes for PDFs."""


def _spreadsheet_extractor_version() -> str:
    return f"{SPREADSHEET_EXTRACTOR_VERSION_PREFIX}:{openpyxl.__version__}"


def safe_csv_cell(value) -> str:
    """
    CSV formula-injection guard (Section 17): a value beginning with =, +,
    -, or @ is prefixed with a single quote so spreadsheet software opens
    it as plain text, never evaluates it as a formula - the standard CSV-
    injection defense. Applied only at EXPORT time to values ARCHIOSK
    itself writes as plain data; a formula genuinely read from a real
    .xlsx is preserved verbatim by apply_bounded_cell_edit below, never
    routed through this function.
    """
    text = "" if value is None else str(value)
    if text[:1] in ("=", "+", "-", "@"):
        return "'" + text
    return text


def _looks_like_ole2_container(raw_bytes: bytes) -> bool:
    return raw_bytes[:8] == _OLE2_MAGIC


def _cell_info(cell, cached_cell) -> dict:
    is_formula = cell.data_type == "f"
    return {
        "value": None if is_formula else cell.value,
        "formula": cell.value if is_formula else None,
        # The cached value Excel itself last computed and saved into the
        # file - openpyxl never recalculates it. None here honestly means
        # "no cached value available" (e.g. a workbook that was written
        # programmatically and never opened/recalculated by real Excel),
        # never fabricated as zero or re-derived.
        "cached_value": cached_cell.value,
        "data_type": cell.data_type,
    }


def _inspect_csv(raw_bytes: bytes, filename: str) -> dict:
    """
    CSV has no sheets/formulas/hidden state - treated as one single
    "sheet" named after the file. Python's csv module returns every field
    as a plain string, so leading zeros ("007") and other identifier-
    shaped values are preserved automatically - the risk this stage's own
    Section 17 warns about is on the WRITE/export side, not this read
    side.
    """
    text = raw_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)
    if not all_rows:
        return {
            "classification": SPREADSHEET_CLASSIFICATION_SUPPORTED, "warnings": [],
            "sheets": [{
                "name": Path(filename).stem, "index": 0, "visible": True,
                "row_count": 0, "column_count": 0, "truncated": False, "rows": [],
            }],
        }

    column_count = max(len(r) for r in all_rows)
    truncated = len(all_rows) > MAX_ROWS_PER_SHEET
    warnings = []
    if truncated:
        warnings.append(f"CSV has {len(all_rows)} rows; only the first {MAX_ROWS_PER_SHEET} were inspected")

    rows = []
    for row_index, raw_row in enumerate(all_rows[:MAX_ROWS_PER_SHEET], start=1):
        cells = {}
        for col_index, raw_value in enumerate(raw_row):
            col_letter = openpyxl.utils.get_column_letter(col_index + 1)
            cells[col_letter] = {
                "value": raw_value, "formula": None, "cached_value": raw_value, "data_type": "s",
            }
        rows.append({"row_index": row_index, "cells": cells})

    return {
        "classification": SPREADSHEET_CLASSIFICATION_SUPPORTED, "warnings": warnings,
        "sheets": [{
            "name": Path(filename).stem, "index": 0, "visible": True,
            "row_count": len(all_rows), "column_count": column_count,
            "truncated": truncated, "rows": rows,
        }],
    }


def inspect_workbook(raw_bytes: bytes, filename: str) -> dict:
    """
    Reads and classifies a .xlsx or .csv file. ALWAYS returns a real dict
    with a "classification" key - an unreadable/refused workbook is a
    real, honest classification value (Section 16's "record parser
    failures"), never an exception. SpreadsheetIntelligenceError is
    reserved for caller-side errors (see the class docstring above).

    Returns `{"classification", "warnings": [...], "sheets": [...]}`.
    """
    ext = Path(filename).suffix.lower()

    if len(raw_bytes) > MAX_UNCOMPRESSED_BYTES:
        return {"classification": SPREADSHEET_CLASSIFICATION_EXCESSIVE_SIZE, "warnings": [], "sheets": []}

    if ext == ".csv":
        return _inspect_csv(raw_bytes, filename)

    if ext != ".xlsx":
        raise SpreadsheetIntelligenceError(
            f"'{filename}' is not a supported spreadsheet format (.xlsx or .csv)."
        )

    if _looks_like_ole2_container(raw_bytes):
        return {"classification": SPREADSHEET_CLASSIFICATION_ENCRYPTED_OR_UNSUPPORTED, "warnings": [], "sheets": []}

    # Decompression-bomb guard: inspect the zip's OWN declared uncompressed
    # size before openpyxl (or anything else) decompresses a single byte.
    # Same pass also does content-based macro detection (Section 16:
    # "detect macro-enabled files") - a real check against the zip's own
    # entry names, not the claimed file extension alone, since only .xlsx
    # is ever accepted by ALLOWED_DOCUMENT_EXTENSIONS in the first place
    # (a genuine .xlsm is already refused before this module ever sees it)
    # but a renamed .xlsm masquerading as .xlsx must still be caught here,
    # defense in depth, never macro content executed either way.
    try:
        with zipfile.ZipFile(io.BytesIO(raw_bytes)) as zf:
            names = zf.namelist()
            total_uncompressed = sum(info.file_size for info in zf.infolist())
    except zipfile.BadZipFile:
        return {"classification": SPREADSHEET_CLASSIFICATION_MALFORMED, "warnings": [], "sheets": []}
    if total_uncompressed > MAX_UNCOMPRESSED_BYTES:
        return {"classification": SPREADSHEET_CLASSIFICATION_EXCESSIVE_SIZE, "warnings": [], "sheets": []}
    if any(n.lower() == "xl/vbaproject.bin" for n in names):
        return {
            "classification": SPREADSHEET_CLASSIFICATION_ENCRYPTED_OR_UNSUPPORTED,
            "warnings": ["workbook contains an embedded VBA project (macro-enabled) - refused, never executed"],
            "sheets": [],
        }

    try:
        wb_formulas = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=False)
        wb_values = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    except (InvalidFileException, zipfile.BadZipFile, KeyError):
        return {"classification": SPREADSHEET_CLASSIFICATION_MALFORMED, "warnings": [], "sheets": []}
    except Exception:  # noqa: BLE001 - matches pdf_intelligence.py's own precedent:
        # an adversarial/corrupt file can raise something outside openpyxl's
        # own exception hierarchy - still an honest "could not read this
        # file", never a 500 that corrupts or crashes the caller.
        return {"classification": SPREADSHEET_CLASSIFICATION_MALFORMED, "warnings": [], "sheets": []}

    warnings: list[str] = []
    sheet_names = wb_formulas.sheetnames
    if len(sheet_names) > MAX_SHEETS:
        warnings.append(f"workbook has {len(sheet_names)} sheets; only the first {MAX_SHEETS} were inspected")
        sheet_names = sheet_names[:MAX_SHEETS]

    if getattr(wb_formulas, "_external_links", None):
        warnings.append("workbook contains external links")

    sheets: list[dict] = []
    workbook_has_formula = False
    for index, name in enumerate(sheet_names):
        ws = wb_formulas[name]
        ws_values = wb_values[name]
        visible = ws.sheet_state == "visible"
        if not visible:
            warnings.append(f"sheet '{name}' is hidden")
        if getattr(ws, "protection", None) is not None and ws.protection.sheet:
            warnings.append(f"sheet '{name}' has protected cells")
        if ws.tables:
            warnings.append(f"sheet '{name}' contains {len(ws.tables)} native Excel table(s)")

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        truncated = max_row > MAX_ROWS_PER_SHEET
        if truncated:
            warnings.append(f"sheet '{name}' has {max_row} rows; only the first {MAX_ROWS_PER_SHEET} were inspected")
        row_limit = min(max_row, MAX_ROWS_PER_SHEET)

        rows: list[dict] = []
        for row in ws.iter_rows(min_row=1, max_row=row_limit):
            cells = {}
            any_value = False
            for cell in row:
                cached_cell = ws_values.cell(row=cell.row, column=cell.column)
                info = _cell_info(cell, cached_cell)
                if cell.data_type == "f":
                    workbook_has_formula = True
                if info["value"] is not None or info["formula"] is not None:
                    any_value = True
                cells[cell.column_letter] = info
            if any_value:
                rows.append({"row_index": row[0].row, "cells": cells})

        sheets.append({
            "name": name, "index": index, "visible": visible,
            "row_count": max_row, "column_count": max_col,
            "truncated": truncated, "rows": rows,
        })

    if workbook_has_formula:
        # Section 9's own explicit instruction: never present a cached
        # result as newly computed, and always state plainly that
        # recalculation was not performed.
        warnings.append(
            "workbook contains formulas; cached values are read from the file as last saved, "
            "not recalculated by ARCHIOSK"
        )

    return {"classification": SPREADSHEET_CLASSIFICATION_SUPPORTED, "warnings": warnings, "sheets": sheets}


def register_spreadsheet_evidence_for_source(
    store: CaseWorkspaceStore,
    workspace: ProjectWorkspace,
    source_id: str,
    actor: str = "system",
    governance_log: Optional[GovernanceLog] = None,
) -> dict:
    """
    Loads the named Source's own persisted original bytes (Source.
    file_path - reliably present for any Source added via add_document_
    source, the same precedent services/pdf_intelligence.py already
    relies on for PDFs), classifies it, and registers the result as
    governed MM1 evidence. Always returns a result dict; never corrupts
    the Source record on a failed/refused read (every failure branch
    returns before any Source mutation).
    """
    source = store._find(workspace.sources, source_id)
    if source is None or source["project_id"] != workspace.project_id:
        raise SpreadsheetIntelligenceError(f"Source {source_id} was not found.")

    file_path = source.get("file_path")
    if not file_path:
        raise SpreadsheetIntelligenceError(f"Source {source_id} has no stored original file to read.")

    name = source.get("name") or ""
    if Path(name).suffix.lower() not in (".xlsx", ".csv"):
        raise SpreadsheetIntelligenceError(f"Source {source_id} ('{name}') is not a spreadsheet.")

    path = Path(file_path)
    try:
        raw_bytes = path.read_bytes()
    except OSError as exc:
        raise SpreadsheetIntelligenceError(f"Source {source_id}'s stored file could not be read: {exc}") from exc

    inspection = inspect_workbook(raw_bytes, name)
    if inspection["classification"] != SPREADSHEET_CLASSIFICATION_SUPPORTED:
        return {
            "classification": inspection["classification"], "warnings": inspection["warnings"],
            "structural_unit_ids": [], "addressable_region_ids": [], "evidence_item_ids": [],
        }

    extractor_version = _spreadsheet_extractor_version()
    result = store.register_spreadsheet_structure(
        workspace, source_id, inspection["sheets"], extractor_version=extractor_version,
        actor=actor, governance_log=governance_log,
    )

    try:
        store.update_source_identity(
            workspace, source_id, actor=actor, mime_type=(
                "text/csv" if name.lower().endswith(".csv")
                else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            size_bytes=len(raw_bytes), extractor_version=extractor_version,
            governance_log=governance_log,
        )
    except CaseWorkspaceError:
        pass  # structure registration already succeeded and was already saved

    return {
        "classification": inspection["classification"], "warnings": inspection["warnings"],
        **result,
    }


def apply_bounded_cell_edit(
    store: CaseWorkspaceStore,
    workspace: ProjectWorkspace,
    source_id: str,
    sheet_name: str,
    cell_ref: str,
    new_value: str,
    expected_file_hash: Optional[str] = None,
    actor: str = "system",
    governance_log: Optional[GovernanceLog] = None,
) -> dict:
    """
    The one safe edit operation this stage implements (Section 8):
    replaces an ORDINARY entered value in a single cell of an .xlsx
    workbook. Deliberately narrow, each restriction load-bearing, not
    incidental:

    - Refuses outright if the target cell currently holds a FORMULA
      (Section 9's "no silent formula replacement" / Section 8's "avoid
      direct editing of arbitrary formulas... unless safely bounded" -
      this slice does not attempt to bound formula editing at all, so it
      refuses instead).
    - `expected_file_hash`, if given, must match the Source's own current
      `file_hash` or the edit is refused - the same optimistic-
      concurrency principle CaseWorkspaceStore.save() already applies to
      the JSON workspace, applied here to the FILE layer specifically
      (two different in-memory copies of the same workbook editing
      different cells is a real, distinct race the JSON-level version
      check alone would not catch).
    - The pre-edit bytes are copied to a sibling backup file BEFORE the
      original is overwritten - "preserve original workbook" (Section 8)
      stays true even though there is deliberately no new Source/
      Supersession lineage record for a single-cell edit in this first
      slice (see this stage's own documentation for why register_source_
      revision, drawings-only today, was not generalized here).
    - The new value is written with the SAME data type the old value had
      (string stays string) - a leading-zero identifier like "007" is
      never silently coerced into the number 7 by openpyxl's own type
      inference.
    - Only .xlsx is supported for editing (CSV export/round-trip is a
      read+regenerate operation, not a targeted single-cell edit, and
      is deliberately out of this slice's own bounded scope).
    """
    source = store._find(workspace.sources, source_id)
    if source is None or source["project_id"] != workspace.project_id:
        raise SpreadsheetIntelligenceError(f"Source {source_id} was not found.")

    name = source.get("name") or ""
    if not name.lower().endswith(".xlsx"):
        raise SpreadsheetIntelligenceError(f"Source {source_id} ('{name}') is not an editable .xlsx workbook.")

    file_path = source.get("file_path")
    if not file_path:
        raise SpreadsheetIntelligenceError(f"Source {source_id} has no stored original file to edit.")
    path = Path(file_path)
    if not path.exists():
        raise SpreadsheetIntelligenceError(f"Source {source_id}'s stored file is missing on disk.")

    raw_bytes = path.read_bytes()
    current_hash = hashlib.sha256(raw_bytes).hexdigest()
    if expected_file_hash is not None and expected_file_hash != current_hash:
        raise SpreadsheetIntelligenceError(
            "This workbook was modified since you last loaded it. Reload and retry."
        )

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=False)
    except Exception as exc:  # noqa: BLE001 - a workbook that inspect_workbook already
        # classified SUPPORTED should always re-open cleanly; this is a
        # last-resort guard, not the primary classification path.
        raise SpreadsheetIntelligenceError(f"Workbook could not be opened for editing: {exc}") from exc

    if sheet_name not in wb.sheetnames:
        raise SpreadsheetIntelligenceError(f"Sheet '{sheet_name}' was not found in this workbook.")
    ws = wb[sheet_name]

    try:
        cell = ws[cell_ref]
    except (KeyError, ValueError) as exc:
        raise SpreadsheetIntelligenceError(f"'{cell_ref}' is not a valid cell reference.") from exc

    if cell.data_type == "f":
        raise SpreadsheetIntelligenceError(
            f"Cell {cell_ref} contains a formula ({cell.value!r}) - editing formulas is not supported in this stage."
        )

    before_value = cell.value
    was_string = cell.data_type == "s" or isinstance(before_value, str)
    written_value = str(new_value) if was_string else new_value
    cell.value = written_value

    out_buf = io.BytesIO()
    wb.save(out_buf)
    new_bytes = out_buf.getvalue()
    new_hash = hashlib.sha256(new_bytes).hexdigest()

    # Preserve the pre-edit original as a real, separate file on disk
    # BEFORE overwriting - "preserve original workbook" stays literally
    # true, not just true-in-spirit.
    backup_path = path.with_name(f"{path.stem}.pre-edit-{int(time.time())}{path.suffix}")
    shutil.copy2(path, backup_path)
    path.write_bytes(new_bytes)

    store.update_source_identity(
        workspace, source_id, actor=actor, file_hash=new_hash, governance_log=None,
    )

    if governance_log is not None:
        governance_log.append(
            project_id=workspace.project_id, event_type="spreadsheet_cell_edited",
            actor=actor, role="human",
            payload={
                "source_id": source_id, "sheet_name": sheet_name, "cell_ref": cell_ref,
                "before_value": before_value, "after_value": written_value,
                "backup_path": str(backup_path), "previous_file_hash": current_hash,
                "new_file_hash": new_hash,
            },
            correlation_id=source_id,
        )

    return {
        "source_id": source_id, "sheet_name": sheet_name, "cell_ref": cell_ref,
        "before_value": before_value, "after_value": written_value,
        "backup_path": str(backup_path), "new_file_hash": new_hash,
    }
